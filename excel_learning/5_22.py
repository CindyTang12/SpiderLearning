import openpyxl

# 创建一个Excel workbook 对象
book = openpyxl.Workbook()

# 创建时，会自动产生一个sheet，通过active获取
sh = book.active

# 修改当前 sheet 标题为 工资表
sh.title = '工资表'

# 保存文件
book.save('信息.xlsx')

# 增加一个名为 '年龄表' 的sheet，放在最后
sh1 = book.create_sheet('年龄表-最后')

# 增加一个 sheet，放在最前
sh2 = book.create_sheet('年龄表-最前', 0)

# 增加一个 sheet，指定为第2个表单
sh3 = book.create_sheet('年龄表2', 1)

# 根据名称获取某个sheet对象
sh = book['工资表']

# 给第一个单元格写入内容
sh['A1'] = '你好'

# 获取某个单元格内容
print(sh['A1'].value)

# 根据行号列号， 给第一个单元格写入内容，
# 注意和 xlrd 不同，是从 1 开始
sh.cell(2, 2).value = '白月黑羽'

# 根据行号列号， 获取某个单元格内容
print(sh.cell(1, 1).value)

book.save('信息.xlsx')
